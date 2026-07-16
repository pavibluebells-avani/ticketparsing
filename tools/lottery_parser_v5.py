"""
lottery_parser_v5.py — Context-Aware Lottery Message Parser V5

Architecture:
  1. NORMALIZER  — cleans separators, unicode, whitespace
  2. LEXICON     — alias dictionary for lottery, rate, bet_type, modifiers
  3. TOKENIZER   — line-by-line token classification
  4. CONTEXT FSM — carries lottery/timeslot/rate/bet_type state across lines
  5. EXTRACTOR   — produces canonical entries from tokens + context
  6. EXPANDER    — box permutations, "all" expansion, range expansion
  7. REPORTER    — Excel output with dashboard

Usage:
  python lottery_parser_v5.py [--input data/raw] [--output reports]
"""

from __future__ import annotations

import json
import re
import os
import sys
import glob
import itertools
import argparse
import logging
from datetime import datetime, timezone, timedelta
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Any, Optional, Tuple
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger("lottery_parser_v5")

# ============================================================
# CONSTANTS
# ============================================================

VERSION = "5.0"

# ============================================================
# LEXICAL DICTIONARY
# ============================================================

LOTTERY_ALIASES = {
    "dear": "DEAR", "dr": "DEAR", "deer": "DEAR", "d": "DEAR",
    "dear.": "DEAR", "dr.": "DEAR",
    "kerala": "KERALA", "kl": "KERALA", "k,l": "KERALA", "k.l": "KERALA",
    "kerla": "KERALA", "keral": "KERALA", "kl.": "KERALA",
    "goa": "GOA", "go": "GOA",
}

TIMESLOT_PATTERNS = [
    (re.compile(r'\b1\s*[.:]*\s*pm\b', re.I), "1PM"),
    (re.compile(r'\bpm\s*1\b', re.I), "1PM"),
    (re.compile(r'\b3\s*[.:]*\s*pm\b', re.I), "3PM"),
    (re.compile(r'\bpm\s*3\b', re.I), "3PM"),
    (re.compile(r'\b6\s*[.:]*\s*pm\b', re.I), "6PM"),
    (re.compile(r'\bpm\s*6\b', re.I), "6PM"),
    (re.compile(r'\b7\s*[.:]*\s*pm\b', re.I), "7PM"),
    (re.compile(r'\b8\s*[.:]*\s*pm\b', re.I), "8PM"),
    (re.compile(r'\b10\s*[.:]*\s*pm\b', re.I), "10PM"),
    (re.compile(r'\bpm\s*10\b', re.I), "10PM"),
]

# Draw schedule per lottery — ordered list of draw hours in IST (24h)
# Used for timestamp-based timeslot fallback when message text doesn't specify
LOTTERY_DRAW_SCHEDULE = {
    "DEAR":   [13, 18, 20],       # 1PM, 6PM, 8PM
    "KERALA": [15],                # 3PM only
    "GOA":    [12, 17, 19],        # 12PM, 5PM, 7PM
}

# Map draw hour → timeslot label
DRAW_HOUR_TO_TIMESLOT = {
    12: "12PM", 13: "1PM", 15: "3PM", 17: "5PM",
    18: "6PM", 19: "7PM", 20: "8PM",
}

IST = timezone(timedelta(hours=5, minutes=30))

BET_TYPE_MAP = {
    "ab": "AB", "a.b": "AB", "a b": "AB",
    "bc": "BC", "b.c": "BC", "b c": "BC",
    "ac": "AC", "a.c": "AC", "a c": "AC",
    "abc": "ABC",
    "all": "ALL",
    "allbot": "ALL", "all board": "ALL", "allboard": "ALL",
    "full board": "ALL", "fullboard": "ALL",
    "board": "ALL", "bort": "ALL",
    "single": "SINGLE",  # context: single-digit bet
}

RATE_PATTERN = re.compile(
    r'(?:rs|re|r\.s|r,s|r\s)[.,\s]*(\d+)', re.I
)
RATE_SUFFIX_PATTERN = re.compile(
    r'(\d+)\s*(?:rs|re)\.?\b', re.I
)

SET_ALIASES = {"set", "sets", "sat", "ser", "sett", "seat", "pcs", "pes"}
EACH_ALIASES = {"each", "ecsh", "ech", "eash", "ea"}
BOX_ALIASES = {"box", "bx"}

# Noise patterns — messages to skip entirely
NOISE_PATTERNS = [
    re.compile(r'^$'),
    re.compile(r'^(entry done|reporting completed|mistakes corrected|ok|done|noted|received|ok done|entry|entries done|report done)$', re.I),
    re.compile(r'^(good morning|good night|hi|hello|thanks|thank you|one|gm|gn|okay|draw|correct|send|sended)$', re.I),
    re.compile(r'^[\U0001f600-\U0001f64f\U0001f300-\U0001f5ff\U0001f680-\U0001f6ff\U0001f900-\U0001f9ff☀-⛿✀-➿\s\U0001f44d]+$'),
    re.compile(r'^[^\x00-\x7F\d]+$'),  # Pure non-ASCII (Tamil, emoji-only, etc.)
    re.compile(r'^degtel', re.I),  # degtelC47A2... admin tokens
    re.compile(r'^(statement|tickets?|winning|digital|degital|sended)\b', re.I),
]

# Single digit position aliases
SINGLE_DIGIT_POSITIONS = {"a": "A", "b": "B", "c": "C"}

# ============================================================
# DATA CLASSES
# ============================================================

@dataclass
class ParseContext:
    """Mutable state carried across lines within a message"""
    lottery: Optional[str] = None
    timeslot: Optional[str] = None
    rate: Optional[int] = None
    bet_type: Optional[str] = None  # AB, BC, AC, ABC, ALL, A, B, C
    bet_types: List[str] = field(default_factory=list)  # Multiple: ["AB", "AC"]
    qty_default: int = 1
    date_str: Optional[str] = None

@dataclass
class CanonicalEntry:
    message_id: str
    number: str
    bet_type: Optional[str] = None
    qty: int = 1
    rate: Optional[int] = None
    category: Optional[str] = None
    lottery: Optional[str] = None
    timeslot: Optional[str] = None
    amount: float = 0
    is_box: bool = False
    raw_line: str = ""
    group_name: str = ""
    sender: str = ""
    push_name: str = ""
    timestamp: int = 0
    human_time: str = ""
    source_file: str = ""
    contest_date: Optional[str] = None  # YYYY-MM-DD — draw date (may differ from message date)

@dataclass
class TokenTrace:
    """Records what a token was interpreted as"""
    original: str
    interpretation: str  # e.g. "LOTTERY:DEAR", "TIMESLOT:1PM", "BET_NUMBER:360", "QTY:5"
    confidence: str = "high"  # high, medium, low

@dataclass
class MessageTrace:
    """Full audit trail for a parsed message"""
    message_id: str = ""
    raw_text: str = ""
    normalized_text: str = ""
    token_traces: List[TokenTrace] = field(default_factory=list)
    entries_count: int = 0
    lottery: Optional[str] = None
    timeslot: Optional[str] = None
    rate: Optional[int] = None
    confidence: str = "high"  # high, medium, low
    group_name: str = ""
    sender: str = ""
    push_name: str = ""
    timestamp: int = 0
    human_time: str = ""

# ============================================================
# NORMALIZER
# ============================================================

def normalize_separators(text: str) -> str:
    """Normalize various separator styles to canonical forms"""
    import unicodedata
    # Unicode NFKC normalization — converts fancy chars like 𝔹𝕔 → Bc
    text = unicodedata.normalize('NFKC', text)
    # Replace multiple dots/commas with single space
    text = re.sub(r'[.]{2,}', ' ', text)
    text = re.sub(r'[,]{2,}', ' ', text)
    # Single dots between numbers: "59.95.91.19" → "59 95 91 19"
    # But preserve patterns like "139.2.sat" (number.qty.set)
    text = re.sub(r'(\d{2,5})\.(\d{2,5})', r'\1 \2', text)
    # Apply again for chained: "a.b.c.d" needs two passes
    text = re.sub(r'(\d{2,5})\.(\d{2,5})', r'\1 \2', text)
    # Preserve number=qty patterns (87=1, 395=2) by converting to dash form
    # before blanket = removal. The tokenizer already handles number-qty.
    text = re.sub(r'(\d{1,5})=(\d{1,2})(?=\s|$|[,|])', r'\1-\2', text)
    # Normalize remaining equals signs used as separators
    text = re.sub(r'={1,}', ' ', text)
    # Underscore as number-qty separator: 866_1 → 866-1
    text = re.sub(r'(\d{1,5})_(\d{1,2})(?=\s|$|[.,|])', r'\1-\2', text)
    # Normalize multiple dashes (but keep single dash for number-qty patterns)
    text = re.sub(r'-{2,}', ' ', text)
    # Remove stray colons and trailing dots on tokens
    text = text.replace(':', ' ')
    # "dear.1" → "dear 1", "dear.8" → "dear 8"
    text = re.sub(r'\b(dear|dr|kl|kerala|goa)\.(\d)', r'\1 \2', text, flags=re.I)
    # "rs30." or "50." trailing dots → "rs30" or "50"
    text = re.sub(r'(\d)\.(\s|$)', r'\1 ', text)
    # "bc." "ac." trailing dots
    text = re.sub(r'\b(ab|bc|ac|abc)\.\s', r'\1 ', text, flags=re.I)
    # "rs," → "rs"
    text = re.sub(r'\b(rs|re),', r'\1', text, flags=re.I)
    # "ea.1" → "each 1" patterns
    text = re.sub(r'\bea\.(\d)', r'each \1', text, flags=re.I)
    # "-1set" → "1set" (only when NOT preceded by a digit — "4-5set" should stay)
    text = re.sub(r'(?<!\d)-(\d+(?:set|sat|ser))', r' \1', text, flags=re.I)
    # "4-5set" "0-5set" → "4 5set" "0 5set" (number-dash-Nset = number + qty)
    text = re.sub(r'(\d)-(\d+(?:set|sat|ser|seat|pcs))', r'\1 \2', text, flags=re.I)
    # "2.set" → "2set"
    text = re.sub(r'(\d)\.(?:set|sat|ser)', r'\1set', text, flags=re.I)
    # Tamil rupee: "10ரூ" → "rs10"
    text = re.sub(r'(\d+)ரூ', r'rs\1', text)
    text = re.sub(r'ரூ(\d+)', r'rs\1', text)
    # "60-rs" → "rs60"
    text = re.sub(r'(\d+)\s*-\s*rs\b', r'rs\1', text, flags=re.I)
    # "8pm." → "8pm"
    text = re.sub(r'(\dpm)\.', r'\1', text, flags=re.I)
    # "8.pm" → "8pm"
    text = re.sub(r'(\d)\.pm', r'\1pm', text, flags=re.I)
    # "rs30pm8" → "rs30 8pm"
    text = re.sub(r'\b(rs\d+)(pm)(\d)\b', r'\1 \3\2', text, flags=re.I)
    # "dear8pm" → "dear 8pm"
    text = re.sub(r'\b(dear|dr|kl|kerala)([\d]+pm)', r'\1 \2', text, flags=re.I)
    # "1sets" → "1set"
    text = re.sub(r'(\d+)sets\b', r'\1set', text, flags=re.I)
    # "10stc" → "10set" (common typo)
    text = re.sub(r'(\d+)stc\b', r'\1set', text, flags=re.I)
    # "borad" / "bort" → "board"
    text = re.sub(r'\bborad\b', 'board', text, flags=re.I)
    text = re.sub(r'\bbort\b', 'board', text, flags=re.I)
    # "board." trailing dot
    text = re.sub(r'\bboard\.', 'board', text, flags=re.I)
    # trailing comma on numbers "30," → "30"
    text = re.sub(r'(\d),(\s|$)', r'\1 ', text)
    # Split compound tokens: "bc1set" → "bc 1set", "25rs.dear" → "25rs dear"
    text = re.sub(r'([a-zA-Z]{2,3})(\d+(?:set|sat|ser|seat))', r'\1 \2', text, flags=re.I)
    # Split "Nrs.lottery": "25rs.dear" → "25rs dear"
    text = re.sub(r'(\d+rs)\.([a-zA-Z])', r'\1 \2', text, flags=re.I)
    # Split "rs-N" → "rs N" for rate extraction
    text = re.sub(r'\b(rs|re)\s*[-]\s*(\d+)', r'\1\2', text, flags=re.I)
    # "p1m" "p3m" "p6m" "p8m" → "1pm" "3pm" "6pm" "8pm"
    text = re.sub(r'\bp(\d{1,2})m\b', r'\1pm', text, flags=re.I)
    # "pm8" "pm6" "pm." → normalize pm+digit
    text = re.sub(r'\bpm\.?(\d{1,2})\b', r'\1pm', text, flags=re.I)
    # Indian rupee notation: "10/-" "30/-" → "rs10" "rs30"
    text = re.sub(r'\b(\d+)/\-', r'rs\1', text)
    # "30.rs" "50.rs" → "rs30" "rs50"
    text = re.sub(r'(\d+)\.rs\b', r'rs\1', text, flags=re.I)
    # "dear.rs.50" → "dear rs50"
    text = re.sub(r'\b(dear|dr|kl|kerala)\.(rs)\.(\d+)', r'\1 \2\3', text, flags=re.I)
    # "rs30p6m" → "rs30 6pm"
    text = re.sub(r'(rs\d+)p(\d{1,2})m\b', r'\1 \2pm', text, flags=re.I)
    # "dear6" "dear10" "kl3" "dr1" → "dear 6pm" etc. (lottery + bare timeslot digit)
    # Must come AFTER dearNpm and dear.N rules
    text = re.sub(r'\b(dear|dr)[\s,]*((?:1[02]?|[35678]))(?!\d|pm|rs|set|sat|\.)', r'\1 \2pm', text, flags=re.I)
    text = re.sub(r'\b(kl|kerala)[\s,]*((?:1[02]?|[35678]))(?!\d|pm|rs|set|sat|\.)', r'\1 \2pm', text, flags=re.I)
    # "abacbc" → "ab ac bc"
    text = re.sub(r'\babacbc\b', 'ab ac bc', text, flags=re.I)
    # "abcd" → "all" (all bet types)
    text = re.sub(r'\babcd\b', 'all', text, flags=re.I)
    # "5seat" → "5set"
    text = re.sub(r'(\d+)seat\b', r'\1set', text, flags=re.I)
    # "pes" → "pcs" (pieces = quantity)
    text = re.sub(r'\bpes\b', 'pcs', text, flags=re.I)
    # "1setdear" "1set.dear" → "1set dear"
    text = re.sub(r'(\d+(?:set|sat|ser))\.*\s*(dear|dr|kl|kerala)', r'\1 \2', text, flags=re.I)
    # "1set." trailing dot on set tokens
    text = re.sub(r'(\d+(?:set|sat|ser))\.', r'\1', text, flags=re.I)
    # "c." "b." "a." standalone → remove trailing dot
    text = re.sub(r'\b([abc])\.(?=\s|$)', r'\1', text, flags=re.I)
    # "pm." trailing dot
    text = re.sub(r'\bpm\.(?=\s|$)', 'pm', text, flags=re.I)
    # "abc0" "abc2" "abc5" "abc6" → "abc 0" "abc 2" etc.
    text = re.sub(r'\b(abc)(\d+)\b', r'\1 \2', text, flags=re.I)
    # "kl100" → "kl rs100" (lottery + rate without rs prefix)
    text = re.sub(r'\b(kl|kerala)(\d{2,3})(?!\d|pm)', r'\1 rs\2', text, flags=re.I)
    # "kl," → "kl" (trailing comma)
    text = re.sub(r'\b(kl|kerala|dear|dr),(?=\s|$)', r'\1', text, flags=re.I)
    # "each-2" → "each 2"
    text = re.sub(r'\beach[-](\d+)', r'each \1', text, flags=re.I)
    # "ecah" → "each"
    text = re.sub(r'\becah\b', 'each', text, flags=re.I)
    # "30₹" "₹30" → "rs30" (₹ symbol)
    text = re.sub(r'(\d+)₹', r'rs\1', text)
    text = re.sub(r'₹(\d+)', r'rs\1', text)
    # "76.1set" → "76 1set" (number.Nset)
    text = re.sub(r'(\d{2,5})\.(\d+(?:set|sat|ser))', r'\1 \2', text, flags=re.I)
    # "-30rs" → "rs30" (leading dash rate)
    text = re.sub(r'-(\d+)rs\b', r'rs\1', text, flags=re.I)
    # "bc-002set" → "bc 002 set" → handled by existing bet_type+num patterns
    text = re.sub(r'\b([abc]{2,3})-(\d+)(set|sat|ser)\b', r'\1 \2 \3', text, flags=re.I)
    # "30rs-805,508" → "rs30 805,508" (rate-dash-numbers: split rate from numbers)
    text = re.sub(r'\b(\d+)rs-(\d)', r'rs\1 \2', text, flags=re.I)
    # Comma-separated betting numbers BEFORE IGNORENUM (so "805,508" → "805 508" not IGNORENUM)
    text = re.sub(r'\b(\d{2,5}),(\d{2,5})\b', r'\1 \2', text)
    # Repeat to handle chains like "805,508,588,885"
    text = re.sub(r'\b(\d{2,5}),(\d{2,5})\b', r'\1 \2', text)
    # Large comma-separated numbers like "50,000" → skip (mark as noise-like)
    # This now only catches true large numbers since betting number commas were already split
    text = re.sub(r'\b\d{1,3},\d{3}\b', 'IGNORENUM', text)
    # Clean up IGNORENUM patterns
    text = re.sub(r'IGNORENUM[,\s]*IGNORENUM', 'IGNORENUM', text)
    # "bcabac" → "bc ab ac" (compound bet types)
    text = re.sub(r'\bbcabac\b', 'bc ab ac', text, flags=re.I)
    text = re.sub(r'\babacbc\b', 'ab ac bc', text, flags=re.I)
    text = re.sub(r'\babbc\b', 'ab bc', text, flags=re.I)
    text = re.sub(r'\bacbc\b', 'ac bc', text, flags=re.I)
    text = re.sub(r'\babac\b', 'ab ac', text, flags=re.I)
    # "(box)" → "box" (parentheses)
    text = text.replace('(box)', 'box').replace('(BOX)', 'box')
    # "bcrs12" → "bc rs12"
    text = re.sub(r'\b([abc]{2,3})(rs\d+)', r'\1 \2', text, flags=re.I)
    # "30rupise" "30rupees" → "rs30"
    text = re.sub(r'(\d+)\s*(?:rupise|rupees|rupee|rupi)\b', r'rs\1', text, flags=re.I)
    # "kl,rs30" → "kl rs30"
    text = re.sub(r'\b(kl|kerala|dear|dr),(rs\d+)', r'\1 \2', text, flags=re.I)
    # "dear/8pm/rs.30" → "dear 8pm rs30" (slash as separator between context tokens)
    # But preserve "123/2" (number/qty) and "123/box" patterns
    text = re.sub(r'([a-zA-Z])/(\d)', r'\1 \2', text)
    text = re.sub(r'(\d)/([a-zA-Z])', r'\1 \2', text)
    text = re.sub(r'([a-zA-Z])/([a-zA-Z])', r'\1 \2', text)
    # "der" → "dear" (typo)
    text = re.sub(r'\bder\b', 'dear', text, flags=re.I)
    # "each." → "each"
    text = re.sub(r'\beach\.', 'each', text, flags=re.I)
    # "pm," → "pm"
    text = re.sub(r'\bpm,', 'pm', text, flags=re.I)
    # "1pmmp" → "1pm" (garbled)
    text = re.sub(r'(\d)pmmp\b', r'\1pm', text, flags=re.I)
    # "all." → "all"
    text = re.sub(r'\ball\.', 'all', text, flags=re.I)
    # "10pic" → "10pcs" (quantity)
    text = re.sub(r'(\d+)pic\b', r'\1pcs', text, flags=re.I)
    # "allbod" → "allboard" → ALL
    text = re.sub(r'\ballbod\b', 'allbot', text, flags=re.I)
    # "abc-00" → "abc rs10" (abc0 variant with dash)
    text = re.sub(r'\babc[-]00?\b', 'abc 0', text, flags=re.I)
    # "dear-1pm" → "dear 1pm"
    text = re.sub(r'\b(dear|dr|kl|kerala)-(\d+pm)', r'\1 \2', text, flags=re.I)
    # "1pm." → "1pm"
    text = re.sub(r'(\dpm)\.', r'\1', text, flags=re.I)
    # "4,set" → "4set"
    text = re.sub(r'(\d+),\s*(?:set|sat|ser)', r'\1set', text, flags=re.I)
    # "30ru" → "rs30" (truncated rupees)
    text = re.sub(r'(\d+)ru\b', r'rs\1', text, flags=re.I)
    # "527x1" "527×1" "527X2" → "527*1" (x/× as multiplication = qty)
    text = re.sub(r'(\d+)[x×](\d+)', r'\1*\2', text, flags=re.I)
    # Standalone "×2" "×1" "×3" after a number token → "*2" etc.
    text = re.sub(r'[×](\d+)', r'*\1', text)
    # "×2nos" "×3nos" → "*2" (nos = numbers/pieces, redundant)
    text = re.sub(r'[×x](\d+)\s*nos\b', r'*\1', text, flags=re.I)
    # "35.1" "53.1" "76.2" → number.qty patterns (2-digit.single = num + qty)
    text = re.sub(r'\b(\d{2,5})\.(\d)(?=\s|$)', r'\1-\2', text)
    # "38/83(5)pcs" → "38 83 5pcs" (parenthesized qty)
    text = re.sub(r'\((\d+)\)(pcs|set|sat|ser)', r' \1\2', text, flags=re.I)
    text = text.replace('(', ' ').replace(')', ' ')
    # "8,pm" "6,pm" → "8pm" "6pm"
    text = re.sub(r'(\d),pm\b', r'\1pm', text, flags=re.I)
    # (comma-separated betting numbers already handled earlier, before IGNORENUM)
    # "1.00pm" "6.00pm" → "1pm" "6pm"
    text = re.sub(r'(\d)\.00pm\b', r'\1pm', text, flags=re.I)
    # "each3" → "each 3"
    text = re.sub(r'\beach(\d)', r'each \1', text, flags=re.I)
    # "b0x" → "box" (zero for o)
    text = re.sub(r'\bb0x\b', 'box', text, flags=re.I)
    # "5st" → "5set" (truncated)
    text = re.sub(r'(\d+)st\b', r'\1set', text, flags=re.I)
    # "set." → "set"
    text = re.sub(r'\bset\.', 'set', text, flags=re.I)
    # "b-9" → "b 9" (bet type B, number 9)
    text = re.sub(r'\b([abc])-(\d)', r'\1 \2', text, flags=re.I)
    # "38/83" "827/287" → "38 83" "827 287" (slash-separated same-length numbers)
    text = re.sub(r'\b(\d{2,5})/(\d{2,5})\b', r'\1 \2', text)
    # Apply again for chains: "55/66/77" needs two passes
    text = re.sub(r'\b(\d{2,5})/(\d{2,5})\b', r'\1 \2', text)
    # Trailing "/." or "/" after number groups
    text = re.sub(r'/\.', ' ', text)
    text = re.sub(r'(\d)/(?=[\s,]|$)', r'\1', text)
    # "dl" → "dear" (another abbreviation)
    text = re.sub(r'\bdl\b', 'dear', text, flags=re.I)
    # "ac-35" "bc-53" → "ac 35" "bc 53" (bet_type-number)
    text = re.sub(r'\b([abc]{2,3})-(\d{1,5})\b', r'\1 \2', text, flags=re.I)
    # "eech" → "each"
    text = re.sub(r'\beech\b', 'each', text, flags=re.I)
    # "1set," → "1set" (trailing comma on set)
    text = re.sub(r'(\d+(?:set|sat|ser)),', r'\1', text, flags=re.I)
    # "1pmrs30" → "1pm rs30"
    text = re.sub(r'(\dpm)(rs\d+)', r'\1 \2', text, flags=re.I)
    # "rs.100.3pm" → "rs100 3pm"
    text = re.sub(r'(rs\.?\d+)\.(\dpm)', r'\1 \2', text, flags=re.I)
    # ".45" ".10set" → "45" "10set" (leading dots)
    text = re.sub(r'(?:^|\s)\.(\d)', r' \1', text)
    # "-30" standalone rate → "rs30"
    text = re.sub(r'(?:^|\s)-(\d+)(?=\s|$)', r' rs\1', text)
    # "ab,bc,ac" "ab-ac" "ac-bc" "ac-bc-ab" → split into separate bet types
    text = re.sub(r'\b(ab|ac|bc)[,\-](ab|ac|bc)(?:[,\-](ab|ac|bc))?', r'\1 \2 \3', text, flags=re.I)
    # "8pmrs.100" → "8pm rs100"
    text = re.sub(r'(\dpm)(rs\.?\d+)', r'\1 \2', text, flags=re.I)
    # "e3set" → "each 3set" (garbled each)
    text = re.sub(r'\be(\d+)(set|sat|ser)', r'each \1\2', text, flags=re.I)
    # "30rupice" → "rs30"
    text = re.sub(r'(\d+)\s*(?:rupice|rupise|rupees|rupee|rupi|ru)\b', r'rs\1', text, flags=re.I)
    # "depr" → "dear" (typo)
    text = re.sub(r'\bdepr\b', 'dear', text, flags=re.I)
    # "k.l.rs.100" → "kl rs100"
    text = re.sub(r'\bk\.l\b', 'kl', text, flags=re.I)
    # "0.5set" → "0 5set" (likely two things: A/B/C number 0, qty 5set)
    # Actually "0.5set" means number with qty — already handled by dot_qty pattern
    # "6.00" → "6" (trailing .00)
    text = re.sub(r'(\d)\.00\b', r'\1', text)
    # "05/27" "72/38" — slash between same-length numbers already handled above
    # Re-apply comma-separated number split (might have been created by earlier rules)
    text = re.sub(r'\b(\d{2,5}),(\d{2,5})\b', r'\1 \2', text)
    # "klrs60" → "kl rs60"
    text = re.sub(r'\b(kl|kerala|dear|dr)(rs\.?\d+)', r'\1 \2', text, flags=re.I)
    # "kl.rs.100" → "kl rs100"
    text = re.sub(r'\b(kl|kerala|dear|dr)\.(rs)\.(\d+)', r'\1 \2\3', text, flags=re.I)
    # "58,85each" → "58 85 each"
    text = re.sub(r'(\d{2,5}),(\d{2,5})(each)', r'\1 \2 \3', text, flags=re.I)
    # "_2set" → "2set"
    text = re.sub(r'_(\d)', r'\1', text)
    # "1,pmrs30" → "1pm rs30"
    text = re.sub(r'(\d),pm', r'\1pm', text, flags=re.I)
    # "00/2set" → "00 2set"
    text = re.sub(r'(\d+)/(\d+(?:set|sat|ser))', r'\1 \2', text, flags=re.I)
    # "92+2set" "69+2set" "72+2set" → number + qty
    text = re.sub(r'(\d+)\+(\d+(?:set|sat|ser))', r'\1 \2', text, flags=re.I)
    # "61+5sed" → "61 5set" (sed typo for set)
    text = re.sub(r'(\d+)sed\b', r'\1set', text, flags=re.I)
    # "8.5set" → "8 5set" (number.Nset)
    text = re.sub(r'\b(\d+)\.(\d+(?:set|sat|ser))', r'\1 \2', text, flags=re.I)
    # "10.se" → "10set" (truncated set)
    text = re.sub(r'(\d+)\.se\b', r'\1set', text, flags=re.I)
    # "0.5set" → "0 5set" (single digit + Nset)
    text = re.sub(r'\b(\d)\.(\d+set)', r'\1 \2', text, flags=re.I)
    # "bm60" → noise (balance/banking message)
    # "5,6" → "5 6" (single digit comma separated)
    text = re.sub(r'\b(\d),(\d)\b', r'\1 \2', text)
    # "ac,80" "ac,bc" "ab," "ac-" "abc-" → strip trailing comma/dash
    text = re.sub(r'\b(ab|ac|bc|abc)[,\-](?=\s|$)', r'\1', text, flags=re.I)
    # "ac,80" → "ac 80" (bet type comma number)
    text = re.sub(r'\b(ab|ac|bc|abc),(\d)', r'\1 \2', text, flags=re.I)
    # "c,8" → "c 8" (single bet type comma number)
    text = re.sub(r'\b([abc]),(\d)', r'\1 \2', text, flags=re.I)
    # "r100" → "rs100" (truncated rs)
    text = re.sub(r'\br(\d{2,3})\b', r'rs\1', text, flags=re.I)
    # "1pmrs30" → "1pm rs30" (reapply if still compound)
    text = re.sub(r'(\dpm)(rs\.?\d+)', r'\1 \2', text, flags=re.I)
    # "61+5set" → "61 5set" (plus-separated)
    text = re.sub(r'(\d+)\+(\d+set)', r'\1 \2', text, flags=re.I)
    # ".set" → "set" (leading dot)
    text = re.sub(r'(?:^|\s)\.set\b', ' set', text, flags=re.I)
    # "058,-1" → "058 -1" (number comma dash number)
    text = re.sub(r'(\d+),(-\d)', r'\1 \2', text)
    # "&" → noise
    text = text.replace('&', ' ')
    # "_" → noise
    text = text.replace('_', ' ')
    # Final cleanup: normalize spaces
    text = re.sub(r'\s+', ' ', text)
    # Normalize multiple spaces
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

# ============================================================
# LOTTERY DETECTION FROM GROUP NAME
# ============================================================

def detect_lottery_from_group(group_name: str) -> Optional[str]:
    gn = group_name.lower()
    if "dear" in gn:
        return "DEAR"
    if "kerala" in gn or "kl" in gn:
        return "KERALA"
    if "goa" in gn:
        return "GOA"
    return None

# ============================================================
# TIMESLOT DETECTION
# ============================================================

def detect_timeslot(text: str) -> Optional[str]:
    for pattern, slot in TIMESLOT_PATTERNS:
        if pattern.search(text):
            return slot
    return None


def resolve_timeslot_from_timestamp(timestamp: int, lottery: Optional[str]) -> Optional[str]:
    """Derive timeslot from WhatsApp timestamp when not explicit in message.

    Uses the lottery's draw schedule to pick the next upcoming draw.
    All comparisons in IST. Consolidator forwards bets before cutoff,
    so the next draw after message time is the intended one.

    Args:
        timestamp: Unix epoch seconds (UTC)
        lottery: Lottery type (DEAR, KERALA, GOA) or None

    Returns:
        Timeslot string like "1PM", "3PM" etc., or None if lottery unknown
    """
    if not timestamp or not lottery:
        return None

    schedule = LOTTERY_DRAW_SCHEDULE.get(lottery)
    if not schedule:
        return None

    # Single-draw lottery (Kerala) — always that timeslot
    if len(schedule) == 1:
        return DRAW_HOUR_TO_TIMESLOT[schedule[0]]

    # Convert to IST
    dt_ist = datetime.fromtimestamp(timestamp, tz=IST)
    hour_ist = dt_ist.hour
    minute_ist = dt_ist.minute
    current_minutes = hour_ist * 60 + minute_ist

    # Find the next draw: first draw hour where draw_hour * 60 > current_minutes
    # (bets arrive before the draw, so next draw is the target)
    for draw_hour in schedule:
        if current_minutes < draw_hour * 60:
            return DRAW_HOUR_TO_TIMESLOT[draw_hour]

    # Past all draws for the day — next day's first draw
    return DRAW_HOUR_TO_TIMESLOT[schedule[0]]


# ============================================================
# RATE EXTRACTION
# ============================================================

def extract_rate(text: str) -> Optional[int]:
    """Extract rate from text like 'Rs.30', 'Rs 100', '60rs', 'Re 60'"""
    m = RATE_PATTERN.search(text)
    if m:
        val = int(m.group(1))
        if val in (10, 12, 15, 20, 25, 30, 50, 60, 100, 650):
            return val
    m = RATE_SUFFIX_PATTERN.search(text)
    if m:
        val = int(m.group(1))
        if val in (10, 12, 15, 20, 25, 30, 50, 60, 100, 650):
            return val
    return None

# ============================================================
# BET TYPE DETECTION
# ============================================================

def detect_bet_type(text: str) -> Optional[str]:
    """Detect bet type header from a line (returns first found)"""
    cleaned = text.lower().strip()
    # Direct match
    if cleaned in BET_TYPE_MAP:
        return BET_TYPE_MAP[cleaned]
    # Check for standalone bet type tokens
    tokens = re.split(r'[\s.,]+', cleaned)
    for t in tokens:
        if t in BET_TYPE_MAP:
            return BET_TYPE_MAP[t]
    return None

def detect_bet_types_multi(text: str) -> List[str]:
    """Detect ALL bet types on a line. 'Ab ac' → ['AB', 'AC']"""
    tokens = re.split(r'[\s.,]+', text.lower().strip())
    found = []
    for t in tokens:
        if t in BET_TYPE_MAP:
            bt = BET_TYPE_MAP[t]
            if bt not in found:
                found.append(bt)
    return found

# ============================================================
# NOISE FILTER
# ============================================================

def is_noise(text: str) -> bool:
    """Check if message is noise (reactions, status updates, etc.)"""
    cleaned = text.strip()
    if not cleaned:
        return True
    for pat in NOISE_PATTERNS:
        if pat.match(cleaned):
            return True
    # Pure emoji messages
    if all(ord(c) > 127 or c.isspace() for c in cleaned):
        return True
    return False

# ============================================================
# NUMBER-QTY EXTRACTION PATTERNS
# ============================================================

# Patterns for number with explicit quantity
RE_NUM_QTY_STAR = re.compile(r'^(\d{1,5})\s*\*\s*(\d+)$')          # 123*2
RE_NUM_QTY_DASH = re.compile(r'^(\d{1,5})\s*-\s*(\d+)$')           # 123-2
RE_NUM_QTY_SLASH = re.compile(r'^(\d{1,5})\s*/\s*(\d+)$')          # 123/2
RE_NUM_QTY_DOT = re.compile(r'^(\d{1,5})\.(\d+)\.(?:sat|set|ser)$', re.I)  # 123.2.sat
RE_NUM_QTY_COMMA = re.compile(r'^(\d{1,5}),(\d+)$')                # 123,2
RE_NUM_QTY_PLUS = re.compile(r'^(\d{1,5})\+(\d+)$')                # 123+1

# Number with box
RE_NUM_BOX = re.compile(r'^(\d{1,5})\s*[/\-.]?\s*box$', re.I)      # 123/box, 123 box

# Standalone number
RE_STANDALONE_NUM = re.compile(r'^(\d{1,5})$')

# BetType.Number.Qty pattern: Ab.84.32 means Ab number=84, number=32 (NOT qty)
# But Bc.02.1set means Bc number=02, qty=1
RE_BETTYPE_NUM_QTY = re.compile(
    r'^([abc]{1,3})\.(\d{1,2})\.(\d+)\.?(?:sat|set|ser)?$', re.I
)

# Rate inline: RS.30.631.845 means Rs30, numbers 631,845
RE_RATE_NUMBERS = re.compile(
    r'^(?:rs|re)\.?(\d+)\.(\d[\d.]+)$', re.I
)

# ============================================================
# CATEGORY DERIVATION
# ============================================================

def derive_category(number: str, rate: Optional[int], bet_type: Optional[str]) -> str:
    """Derive ticket category from number length, rate, and bet_type"""
    dlen = len(number)

    # Single digit positions (A, B, C)
    if dlen == 1:
        return "SINGLE_DIGIT"

    # Two digit — these are AB/BC/AC combination bets
    if dlen == 2:
        return "DOUBLE"

    # Three digit — aligned with dashboard resolveCategory names
    if dlen == 3:
        if rate:
            if rate == 10:
                return "3D10"
            elif rate == 25:
                return "3D25"
            elif rate == 30:
                return "3D_HALF"
            elif rate == 60:
                return "FULL"
            else:
                return f"3D_RS{rate}"
        return "3D_FULL"  # default for 3-digit

    # Four digit — aligned with dashboard resolveCategory names
    if dlen == 4:
        if rate:
            if rate == 20:
                return "4D20"
            elif rate == 50:
                return "4D50"
            elif rate == 100:
                return "4D100"
            else:
                return f"4D_RS{rate}"
        return "4D_RS100"  # default for 4-digit

    if dlen == 5:
        return "5D_RS650"

    return "UNKNOWN"

def get_default_rate(number: str) -> int:
    """Default rate based on digit count"""
    dlen = len(number)
    defaults = {1: 12, 2: 12, 3: 60, 4: 100, 5: 650}
    return defaults.get(dlen, 0)

# Valid rates per digit count — used to check if a carried rate applies
VALID_RATES_BY_DIGITS = {
    1: {12},
    2: {12},                          # AB/BC/AC always Rs12
    3: {10, 25, 30, 60},              # 3D10, 3D25, 3D_HALF, FULL
    4: {10, 20, 25, 50, 100},         # 4D20, 4D50, 4D100
    5: {650},                         # 5D
}

def validate_rate_for_digits(rate: int, number: str) -> int:
    """Check if carried rate is valid for this digit count.
    If not, return the default rate for that digit count."""
    dlen = len(number)
    valid = VALID_RATES_BY_DIGITS.get(dlen, set())
    if rate in valid:
        return rate
    return get_default_rate(number)

# ============================================================
# BOX EXPANSION
# ============================================================

def expand_box(number: str) -> List[str]:
    """Generate all unique permutations of a number"""
    perms = set(''.join(p) for p in itertools.permutations(number))
    return sorted(perms)

# ============================================================
# RANGE EXPANSION
# ============================================================

def expand_range(start: int, end: int, width: int) -> List[str]:
    """Expand 'X to Y' into list of numbers"""
    result = []
    for n in range(start, end + 1):
        result.append(str(n).zfill(width))
    return result

# ============================================================
# LINE PARSER
# ============================================================

def parse_line(
    line: str,
    ctx: ParseContext,
    message_id: str,
    meta: dict,
    stats: dict,
    traces: Optional[List[TokenTrace]] = None
) -> List[CanonicalEntry]:
    """Parse a single line within a message, updating context as needed."""
    entries = []
    original_line = line
    line = normalize_separators(line)
    line_lower = line.lower().strip()

    if traces is None:
        traces = []

    def trace(orig, interp, conf="high"):
        traces.append(TokenTrace(original=orig, interpretation=interp, confidence=conf))

    if not line_lower:
        return entries

    # --- Detect and update context from this line ---

    # Lottery detection
    for alias, canonical in LOTTERY_ALIASES.items():
        # Match as word boundary
        if re.search(r'(?:^|[\s.,])' + re.escape(alias) + r'(?:[\s.,]|$)', line_lower):
            ctx.lottery = canonical
            break

    # Timeslot detection
    ts = detect_timeslot(line)
    if ts:
        ctx.timeslot = ts

    # Rate detection
    rate = extract_rate(line)
    if rate:
        ctx.rate = rate

    # Check if line is purely a context-setter (lottery/timeslot/rate/date header)
    # e.g. "KL 3pm", "Rs 30", "Dear. 1", "30...5...2026"
    stripped = re.sub(r'[^a-zA-Z0-9]', '', line_lower)

    # Date line: "30...5...2026", "17..7..2026", "17/7/2026"
    date_match = re.match(r'^\s*(\d{1,2})\s*[./\s-]+(\d{1,2})\s*[./\s-]+(\d{4})\s*$', line)
    if date_match:
        dd, mm, yyyy = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
        if 1 <= dd <= 31 and 1 <= mm <= 12 and 2020 <= yyyy <= 2030:
            ctx.date_str = f"{yyyy:04d}-{mm:02d}-{dd:02d}"
            trace(line.strip(), f"DATE:{ctx.date_str}")
        return entries

    # Bet type detection — check if line is a bet type header
    # Lines like "AB", "BC", "AC", "ABC", "All", "Ab Bc Ac", "Ab ac"
    multi_bts = detect_bet_types_multi(line)
    bt = multi_bts[0] if multi_bts else None
    if bt:
        # Check if this line is JUST a bet type (no numbers)
        nums_in_line = re.findall(r'\b\d{1,5}\b', line)
        # Filter out rate numbers
        non_rate_nums = [n for n in nums_in_line if int(n) not in (10, 12, 15, 20, 25, 30, 50, 60, 100, 650)]
        if not non_rate_nums:
            ctx.bet_type = bt
            # Store multiple bet types for expansion (e.g. "Ab ac" → ["AB", "AC"])
            ctx.bet_types = multi_bts if len(multi_bts) > 1 else []
            return entries
        # If bet type + numbers on same line, set context and continue to parse numbers
        ctx.bet_type = bt
        ctx.bet_types = multi_bts if len(multi_bts) > 1 else []

    # Check for single-position headers: "A", "B", "C" with possible digit
    single_pos_match = re.match(r'^([abc])\s*$', line_lower)
    if single_pos_match:
        ctx.bet_type = single_pos_match.group(1).upper()
        return entries

    # Detect "each N set" / "N set" at end of line or standalone
    # Also strip the matched pattern from the line so qty tokens don't become bet numbers
    each_qty = None
    each_strip_pattern = None
    for ea in EACH_ALIASES:
        m = re.search(rf'\b{ea}\s*(\d+)\s*(?:set|sat|ser|seat|pcs|pes)?\b', line_lower)
        if m:
            each_qty = int(m.group(1))
            each_strip_pattern = m
            break
    if each_qty is None:
        set_qty_match = re.search(r'\b(\d+)\s*(?:set|sat|ser|seat|pcs|pes)\s*(?:each)?\b', line_lower)
        if set_qty_match:
            each_qty = int(set_qty_match.group(1))
            each_strip_pattern = set_qty_match

    # Check for "Each -Nset" or "Each N" patterns
    if each_qty is None:
        m = re.search(r'\beach\s*[-]?\s*(\d+)', line_lower)
        if m:
            each_qty = int(m.group(1))
            each_strip_pattern = m

    # Strip the each/set qty pattern from the line to prevent qty number being parsed as bet number
    if each_strip_pattern:
        stripped_text = each_strip_pattern.group(0).strip()
        trace(stripped_text, f"QTY:{each_qty}")
        start, end = each_strip_pattern.start(), each_strip_pattern.end()
        line = line[:start] + ' ' + line[end:]
        line_lower = line.lower()

    # --- Extract number entries from line ---

    # Check for inline rate+numbers pattern: "RS.30.631.845.467" or "Rs.10.428"
    rate_nums_match = RE_RATE_NUMBERS.match(line.strip())
    if rate_nums_match:
        ctx.rate = int(rate_nums_match.group(1))
        nums_str = rate_nums_match.group(2)
        numbers = [n for n in re.split(r'[.\s]+', nums_str) if n and re.match(r'^\d+$', n)]
        for num in numbers:
            if len(num) >= 2:  # skip single digits that might be qty
                effective_rate = validate_rate_for_digits(ctx.rate, num) if ctx.rate else get_default_rate(num)
                effective_qty = each_qty or 1
                entry = CanonicalEntry(
                    message_id=message_id,
                    number=num,
                    bet_type=ctx.bet_type,
                    qty=effective_qty,
                    rate=effective_rate,
                    category=derive_category(num, effective_rate, ctx.bet_type),
                    lottery=ctx.lottery,
                    timeslot=ctx.timeslot,
                    amount=effective_rate * effective_qty,
                    raw_line=original_line,
                    **meta
                )
                entries.append(entry)
                stats["extracted"] += 1
                trace(num, f"BET_NUMBER:{num} QTY:{effective_qty} (inline rate+nums)")
        return entries

    # Tokenize line by common separators
    tokens = re.split(r'[\s]+', line.strip())

    # --- Pre-scan: build next_rate[] so each number gets the rate from its
    #     nearest FOLLOWING rate token.  In "866-1 784-1 30rs 3277-1 20rs",
    #     positions 0-3 get 30, positions 4-5 get 20.
    _rate_re1 = re.compile(r'^(?:rs|re)\.?\d+$', re.I)
    _rate_re2 = re.compile(r'^\d+(?:rs|re)\.?$', re.I)
    next_rate = [None] * len(tokens)
    _running_rate = None
    for _j in range(len(tokens) - 1, -1, -1):
        _tl = tokens[_j].strip().lower()
        if _rate_re1.match(_tl) or _rate_re2.match(_tl):
            _r = extract_rate(tokens[_j])
            if _r:
                _running_rate = _r
        next_rate[_j] = _running_rate

    def _effective_rate_at(pos, num):
        """Pick rate for number at token position: next_rate > ctx.rate > default."""
        r = next_rate[pos] or ctx.rate
        if r:
            return validate_rate_for_digits(r, num)
        return get_default_rate(num)

    i = 0
    entries_before = len(entries)
    while i < len(tokens):
        tok = tokens[i].strip()
        tok_lower = tok.lower()

        if not tok or not tok_lower:
            i += 1
            continue

        # Skip pure non-ASCII tokens (Tamil, emoji, etc.)
        if all(ord(c) > 127 or not c.strip() for c in tok):
            trace(tok, "NOISE:non-ascii")
            i += 1
            continue

        # Skip IGNORENUM patterns (large numbers like 50,000)
        if 'ignorenum' in tok_lower:
            trace(tok, "NOISE:large_number")
            i += 1
            continue

        # Skip pure context tokens already processed
        if tok_lower in LOTTERY_ALIASES or tok_lower in ('rs', 're', 'pm', 'rs.', 're.'):
            if tok_lower in LOTTERY_ALIASES:
                trace(tok, f"LOTTERY:{LOTTERY_ALIASES[tok_lower]}")
            elif tok_lower.startswith(('rs', 're')):
                trace(tok, f"RATE_PREFIX")
            else:
                trace(tok, f"TIMESLOT_KEYWORD")
            i += 1
            continue

        # Bet type tokens appearing inline — set context and continue
        if tok_lower in BET_TYPE_MAP:
            ctx.bet_type = BET_TYPE_MAP[tok_lower]
            trace(tok, f"BET_TYPE:{ctx.bet_type}")
            i += 1
            continue

        # Single position tokens (a, b, c standalone)
        if tok_lower in SINGLE_DIGIT_POSITIONS:
            ctx.bet_type = SINGLE_DIGIT_POSITIONS[tok_lower]
            trace(tok, f"BET_TYPE:{ctx.bet_type}")
            i += 1
            continue

        # "4d" = 4-digit context (informational, skip)
        if re.match(r'^\d+d$', tok_lower):
            trace(tok, f"CONTEXT:{tok_lower}")
            i += 1
            continue

        # Skip known noise/modifier tokens
        if tok_lower in SET_ALIASES | EACH_ALIASES | BOX_ALIASES | {
            'board', 'chance', 'quantity', 'entered', 'wrongly',
            'entry', 'done', 'ok', 'one', 'reporting', 'completed',
            'mistakes', 'corrected', 'sir', 'pls', 'please', '/-',
            '.', 'dear.', 'pm', 'full', 'half', 'super', 'doubles',
            'ji', 'va', 'jiii', 'missing', 'port', 'eh', 'but',
            'mentioned', 'report', 'not', 'the', 'is', 'it', 'in',
            'and', 'for', 'with', 'that', 'this', 'was', 'are',
            'be', 'has', 'have', 'had', 'at', 'by', 'no', 'yes',
            'me', 'my', 'we', 'you', 'he', 'she', 'its', 'our',
            'your', 'their', 'am', 'an', 'or', 'if', 'so', 'do',
            'up', 'on', 'of', 'only', 'just', 'also', 'very',
            'total', 'balance', 'pending', 'check', 'waiting',
            'cancel', 'actual', 'company', 'booking', 'mistake',
            'amount', 'payment', 'ticket', 'tickets', 'five', 'two', 'three',
            'four', 'six', 'seven', 'eight', 'nine', 'ten',
            'number', 'numbers', 'msg', 'message', 'received',
            ',', 'will', 'can', 'need',
            'wrong', 'change', 'instead', 'correction',
            'same', 'above', 'below', 'here', 'there',
            'bro', 'anna', 'enna', 'da', 'pa', 'ya', 'la',
            'please', 'poda', 'seri', 'illa', 'oru', 'ipo',
            'than', 'nee', 'nan', 'enga', 'inga', 'anga',
            'draw', 'correct', 'send', 'sended', 'statement',
            'winning', 'okay', 'degital', 'digital', 'degtel',
            'aii', 'bm', 'digit', '*', 'ignorenum',
            'ifsc', 'value', 'account', 'wait', 'double',
            'win', 'extra', 'enter', 'code', 'der',
            '-2', '-1', '-3', '-5', 'e', 'jii', 'bank',
            'cport', 'ail', 'dally',
            'nos', 'sent', 'bm60', 't', 'test', 'less',
            'miss', '&',
        }:
            trace(tok, "NOISE:skip_word")
            i += 1
            continue

        # Skip set/qty compound tokens: "1set", "2sat", "5ser", "10set", "5seat", "2pcs"
        if re.match(r'^\d+\s*(?:set|sat|ser|sett|seat|pcs|pes)$', tok_lower):
            trace(tok, f"QTY:{tok_lower}")
            i += 1
            continue

        # Skip "to" keyword (used in ranges)
        if tok_lower == 'to':
            trace(tok, "NOISE:keyword")
            i += 1
            continue

        # Standalone "*N" qty modifier — applies to last entry
        star_qty_match = re.match(r'^\*(\d+)(?:nos)?$', tok_lower)
        if star_qty_match:
            qty_val = int(star_qty_match.group(1))
            if entries and qty_val <= 20:
                entries[-1].qty = qty_val
                entries[-1].amount = (entries[-1].rate or 0) * qty_val
            trace(tok, f"QTY:{qty_val}")
            i += 1
            continue

        # "nos" standalone (numbers/pieces) — noise
        if tok_lower in ('nos', 'no.s', 'no'):
            trace(tok, "NOISE:skip_word")
            i += 1
            continue

        # Number*qty: 57*2
        m = RE_NUM_QTY_STAR.match(tok)
        if m:
            num, qty = m.group(1), int(m.group(2))
            effective_rate = _effective_rate_at(i, num)
            entries.append(CanonicalEntry(
                message_id=message_id, number=num,
                bet_type=ctx.bet_type, qty=qty,
                rate=effective_rate,
                category=derive_category(num, effective_rate, ctx.bet_type),
                lottery=ctx.lottery, timeslot=ctx.timeslot,
                amount=effective_rate * qty,
                raw_line=original_line, **meta
            ))
            stats["extracted"] += 1
            trace(tok, f"BET_NUMBER:{num} QTY:{qty}")
            i += 1
            continue

        # Number-qty: 364-2 (but not if it's a range like "61 to 67")
        m = RE_NUM_QTY_DASH.match(tok)
        if m:
            num, qty_str = m.group(1), m.group(2)
            qty = int(qty_str)
            # Same-length numbers separated by dash = two betting numbers (e.g. 56-65, 52-25)
            if len(num) >= 2 and len(qty_str) == len(num) and qty > 20:
                for n in [num, qty_str]:
                    effective_rate = _effective_rate_at(i, n)
                    entries.append(CanonicalEntry(
                        message_id=message_id, number=n,
                        bet_type=ctx.bet_type, qty=each_qty or 1,
                        rate=effective_rate,
                        category=derive_category(n, effective_rate, ctx.bet_type),
                        lottery=ctx.lottery, timeslot=ctx.timeslot,
                        amount=effective_rate * (each_qty or 1),
                        raw_line=original_line, **meta
                    ))
                    stats["extracted"] += 1
                trace(tok, f"BET_NUMBER:{num} BET_NUMBER:{qty_str} (dash-split)")
                i += 1
                continue
            # Heuristic: if qty > 20, it's likely two separate numbers, not a qty
            if qty <= 20:
                effective_rate = _effective_rate_at(i, num)
                entries.append(CanonicalEntry(
                    message_id=message_id, number=num,
                    bet_type=ctx.bet_type, qty=qty,
                    rate=effective_rate,
                    category=derive_category(num, effective_rate, ctx.bet_type),
                    lottery=ctx.lottery, timeslot=ctx.timeslot,
                    amount=effective_rate * qty,
                    raw_line=original_line, **meta
                ))
                stats["extracted"] += 1
                trace(tok, f"BET_NUMBER:{num} QTY:{qty}")
                i += 1
                continue

        # Number/qty: 098/2
        m = RE_NUM_QTY_SLASH.match(tok)
        if m:
            num, qty = m.group(1), int(m.group(2))
            if qty <= 20:
                effective_rate = _effective_rate_at(i, num)
                entries.append(CanonicalEntry(
                    message_id=message_id, number=num,
                    bet_type=ctx.bet_type, qty=qty,
                    rate=effective_rate,
                    category=derive_category(num, effective_rate, ctx.bet_type),
                    lottery=ctx.lottery, timeslot=ctx.timeslot,
                    amount=effective_rate * qty,
                    raw_line=original_line, **meta
                ))
                stats["extracted"] += 1
                trace(tok, f"BET_NUMBER:{num} QTY:{qty}")
                i += 1
                continue

        # Number.qty.set: 139.2.sat
        m = RE_NUM_QTY_DOT.match(tok)
        if m:
            num, qty = m.group(1), int(m.group(2))
            effective_rate = _effective_rate_at(i, num)
            entries.append(CanonicalEntry(
                message_id=message_id, number=num,
                bet_type=ctx.bet_type, qty=qty,
                rate=effective_rate,
                category=derive_category(num, effective_rate, ctx.bet_type),
                lottery=ctx.lottery, timeslot=ctx.timeslot,
                amount=effective_rate * qty,
                raw_line=original_line, **meta
            ))
            stats["extracted"] += 1
            trace(tok, f"BET_NUMBER:{num} QTY:{qty} (dot-set)")
            i += 1
            continue

        # Number+qty: 112+1
        m = RE_NUM_QTY_PLUS.match(tok)
        if m:
            num, qty = m.group(1), int(m.group(2))
            if qty <= 20:
                effective_rate = _effective_rate_at(i, num)
                entries.append(CanonicalEntry(
                    message_id=message_id, number=num,
                    bet_type=ctx.bet_type, qty=qty,
                    rate=effective_rate,
                    category=derive_category(num, effective_rate, ctx.bet_type),
                    lottery=ctx.lottery, timeslot=ctx.timeslot,
                    amount=effective_rate * qty,
                    raw_line=original_line, **meta
                ))
                stats["extracted"] += 1
                trace(tok, f"BET_NUMBER:{num} QTY:{qty} (plus)")
                i += 1
                continue

        # Number/box: 962=box, 123/box
        m = RE_NUM_BOX.match(tok)
        if m:
            num = m.group(1)
            effective_rate = _effective_rate_at(i, num)
            box_nums = expand_box(num)
            for bn in box_nums:
                entries.append(CanonicalEntry(
                    message_id=message_id, number=bn,
                    bet_type=ctx.bet_type, qty=each_qty or 1,
                    rate=effective_rate,
                    category=derive_category(bn, effective_rate, ctx.bet_type),
                    lottery=ctx.lottery, timeslot=ctx.timeslot,
                    amount=effective_rate * (each_qty or 1),
                    is_box=True,
                    raw_line=original_line, **meta
                ))
                stats["extracted"] += 1
            trace(tok, f"BOX_EXPAND:{num}→{','.join(box_nums)}")
            i += 1
            continue

        # Check for "box" as next token
        if RE_STANDALONE_NUM.match(tok) and i + 1 < len(tokens) and tokens[i+1].lower().rstrip('.,') in BOX_ALIASES:
            num = tok
            effective_rate = _effective_rate_at(i, num)
            box_nums = expand_box(num)
            for bn in box_nums:
                entries.append(CanonicalEntry(
                    message_id=message_id, number=bn,
                    bet_type=ctx.bet_type, qty=each_qty or 1,
                    rate=effective_rate,
                    category=derive_category(bn, effective_rate, ctx.bet_type),
                    lottery=ctx.lottery, timeslot=ctx.timeslot,
                    amount=effective_rate * (each_qty or 1),
                    is_box=True,
                    raw_line=original_line, **meta
                ))
                stats["extracted"] += 1
            trace(f"{tok} {tokens[i+1]}", f"BOX_EXPAND:{num}→{','.join(box_nums)}")
            i += 2
            continue

        # BetType patterns: "Bc.47.24.86.45" or "Ab.84.32.54.62"
        # These contain bet_type followed by multiple numbers
        bt_multi_match = re.match(r'^([abc]{1,3})\.(.+)$', tok, re.I)
        if bt_multi_match:
            bt_str = bt_multi_match.group(1).upper()
            rest = bt_multi_match.group(2)
            nums = [n for n in re.split(r'[.\s]+', rest) if n and re.match(r'^\d+$', n)]
            if nums:
                if bt_str in BET_TYPE_MAP.values() or bt_str in SINGLE_DIGIT_POSITIONS.values() or len(bt_str) <= 3:
                    for num in nums:
                        effective_rate = _effective_rate_at(i, num)
                        entries.append(CanonicalEntry(
                            message_id=message_id, number=num,
                            bet_type=bt_str, qty=each_qty or 1,
                            rate=effective_rate,
                            category=derive_category(num, effective_rate, bt_str),
                            lottery=ctx.lottery, timeslot=ctx.timeslot,
                            amount=effective_rate * (each_qty or 1),
                            raw_line=original_line, **meta
                        ))
                        stats["extracted"] += 1
                    trace(tok, f"BETTYPE:{bt_str} NUMBERS:{','.join(nums)}")
                    i += 1
                    continue

        # Rate token: rs30, 60rs, rs.100
        if re.match(r'^(?:rs|re)\.?\d+$', tok_lower) or re.match(r'^\d+(?:rs|re)\.?$', tok_lower):
            inline_rate = extract_rate(tok)
            if inline_rate:
                ctx.rate = inline_rate
            trace(tok, f"RATE:Rs{ctx.rate}")
            i += 1
            continue

        # Timeslot token
        if re.match(r'^\d+pm\d?$', tok_lower) or tok_lower == 'pm':
            trace(tok, f"TIMESLOT:{ctx.timeslot}")
            i += 1
            continue

        # Check for bet type token with number: "Ac40", "Ab43"
        bt_num_match = re.match(r'^([abc]{2})(\d+)$', tok_lower)
        if bt_num_match:
            bt_str = bt_num_match.group(1).upper()
            num = bt_num_match.group(2)
            # Check if next token is qty: "--70set"
            qty = each_qty or 1
            if i + 1 < len(tokens):
                nxt = tokens[i+1].lower().strip()
                qty_m = re.match(r'^(\d+)\s*(?:set|sat|ser)$', nxt)
                if qty_m:
                    qty = int(qty_m.group(1))
                    i += 1
            effective_rate = _effective_rate_at(i, num)
            entries.append(CanonicalEntry(
                message_id=message_id, number=num,
                bet_type=bt_str, qty=qty,
                rate=effective_rate,
                category=derive_category(num, effective_rate, bt_str),
                lottery=ctx.lottery, timeslot=ctx.timeslot,
                amount=effective_rate * qty,
                raw_line=original_line, **meta
            ))
            stats["extracted"] += 1
            trace(tok, f"BETTYPE:{bt_str} BET_NUMBER:{num} QTY:{qty}")
            i += 1
            continue

        # Single position + number: "A4", "B3", "C0"
        sp_match = re.match(r'^([abc])(\d+)$', tok_lower)
        if sp_match:
            bt_str = sp_match.group(1).upper()
            num = sp_match.group(2)
            qty = each_qty or 1
            if i + 1 < len(tokens):
                nxt = tokens[i+1].lower().strip()
                qty_m = re.match(r'^(\d+)\s*(?:set|sat|ser)$', nxt)
                if qty_m:
                    qty = int(qty_m.group(1))
                    i += 1
            effective_rate = _effective_rate_at(i, num)
            entries.append(CanonicalEntry(
                message_id=message_id, number=num,
                bet_type=bt_str, qty=qty,
                rate=effective_rate,
                category=derive_category(num, effective_rate, bt_str),
                lottery=ctx.lottery, timeslot=ctx.timeslot,
                amount=effective_rate * qty,
                raw_line=original_line, **meta
            ))
            stats["extracted"] += 1
            trace(tok, f"POSITION:{bt_str} BET_NUMBER:{num} QTY:{qty}")
            i += 1
            continue

        # Standalone number
        m = RE_STANDALONE_NUM.match(tok)
        if m:
            num = m.group(1)
            num_val = int(num)

            # "abc0" shortcut: ABC bet_type with 0 means Rs10 rate
            if num == '0' and ctx.bet_type == 'ABC':
                ctx.rate = 10
                i += 1
                continue

            # Skip if this looks like a rate, year, or context number
            # BUT allow if we have an active bet_type (the number IS the bet number)
            if not ctx.bet_type and num_val in (10, 12, 15, 20, 25, 30, 50, 60, 100, 650, 2026, 2025):
                # If it's a valid rate value, treat as implicit rate (e.g. "Kl 100 8074")
                if num_val in (10, 12, 15, 20, 25, 30, 50, 60, 100, 650):
                    ctx.rate = num_val
                    trace(tok, f"RATE:Rs{num_val}")
                else:
                    trace(tok, f"CONTEXT:{num_val}")
                i += 1
                continue
            # Skip single-digit numbers if no bet_type context (likely noise)
            if len(num) == 1 and not ctx.bet_type:
                # Could be timeslot "1" for 1PM
                if num in ('1', '3', '6', '7', '8'):
                    # Check if it's a standalone timeslot number
                    i += 1
                    continue
                i += 1
                continue

            # Check if next token provides qty
            qty = each_qty or 1
            if i + 1 < len(tokens):
                nxt = tokens[i+1].lower().strip()
                # Check for qty suffix: "1set", "2sat"
                qty_m = re.match(r'^(\d+)\s*(?:set|sat|ser)$', nxt)
                if qty_m:
                    qty = int(qty_m.group(1))
                    i += 1

            effective_rate = _effective_rate_at(i, num)
            entries.append(CanonicalEntry(
                message_id=message_id, number=num,
                bet_type=ctx.bet_type, qty=qty,
                rate=effective_rate,
                category=derive_category(num, effective_rate, ctx.bet_type),
                lottery=ctx.lottery, timeslot=ctx.timeslot,
                amount=effective_rate * qty,
                raw_line=original_line, **meta
            ))
            stats["extracted"] += 1
            bt_label = f" {ctx.bet_type}" if ctx.bet_type else ""
            trace(tok, f"BET_NUMBER:{num}{bt_label} QTY:{qty}")
            i += 1
            continue

        # Handle range: "61 to 67" (already split into tokens)
        # This is handled by looking at "NUMBER to NUMBER" sequence
        if tok_lower == 'to' or tok_lower == '-':
            i += 1
            continue

        # Unknown token
        stats["unknown_tokens"][tok_lower] = stats["unknown_tokens"].get(tok_lower, 0) + 1
        trace(tok, f"UNKNOWN:{tok_lower}", "low")
        i += 1

    # Expand multiple bet types: "Ab ac" → duplicate each entry for AB and AC
    if ctx.bet_types and len(ctx.bet_types) > 1:
        multi_expanded = []
        for entry in entries:
            # If entry uses any of the multi bet_types, expand to all of them
            if entry.bet_type in ctx.bet_types:
                for bt in ctx.bet_types:
                    e = CanonicalEntry(**{**asdict(entry), 'bet_type': bt})
                    multi_expanded.append(e)
            else:
                multi_expanded.append(entry)
        entries = multi_expanded

    # Apply "ALL" expansion if bet_type is ALL
    expanded = []
    for entry in entries:
        if entry.bet_type == "ALL":
            dlen = len(entry.number)
            if dlen == 1:
                for bt in ["A", "B", "C"]:
                    e = CanonicalEntry(**{**asdict(entry), 'bet_type': bt})
                    expanded.append(e)
            elif dlen == 2:
                for bt in ["AB", "AC", "BC"]:
                    e = CanonicalEntry(**{**asdict(entry), 'bet_type': bt})
                    expanded.append(e)
            else:
                expanded.append(entry)
        elif entry.bet_type == "ABC":
            dlen = len(entry.number)
            if dlen == 2:
                for bt in ["AB", "AC", "BC"]:
                    e = CanonicalEntry(**{**asdict(entry), 'bet_type': bt})
                    expanded.append(e)
            elif dlen == 1:
                for bt in ["A", "B", "C"]:
                    e = CanonicalEntry(**{**asdict(entry), 'bet_type': bt})
                    expanded.append(e)
            else:
                expanded.append(entry)
        else:
            expanded.append(entry)

    return expanded

# ============================================================
# MESSAGE PARSER
# ============================================================

def parse_message(message: dict, stats: dict, prev_timeslot: Optional[str] = None, prev_group: Optional[str] = None) -> Tuple[List[CanonicalEntry], Optional[str]]:
    """Parse a complete WhatsApp message (potentially multi-line).
    Returns (entries, timeslot_used) — timeslot_used for cross-message inheritance."""
    text = message.get("text", "")
    if not text or is_noise(text):
        stats["skipped_noise"] += 1
        return [], prev_timeslot

    message_id = message.get("message_id", "")
    group_name = message.get("group_name", "")
    timestamp = int(message.get("whatsapp_timestamp", 0))

    # Initialize context from group name
    ctx = ParseContext()
    ctx.lottery = detect_lottery_from_group(group_name)

    # Inherit timeslot from previous message in same group if not set
    if prev_timeslot and prev_group == group_name:
        ctx.timeslot = prev_timeslot

    # Common metadata for all entries
    meta = {
        "group_name": group_name,
        "sender": message.get("sender", ""),
        "push_name": message.get("push_name", ""),
        "timestamp": timestamp,
        "human_time": datetime.fromtimestamp(timestamp, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if timestamp else "",
        "source_file": message.get("_source_file", ""),
    }

    # Pre-scan full text for rate if it appears at the end (common pattern)
    # e.g. "281\n182\n678\nRs.10" — rate declared at end applies to all
    lines = text.split('\n')
    full_text_lower = text.lower()

    # Check if rate appears ONLY at end (last non-empty line)
    non_empty_lines = [l.strip() for l in lines if l.strip()]
    if non_empty_lines:
        last_line = non_empty_lines[-1]
        last_rate = extract_rate(last_line)
        # If last line is ONLY a rate (no numbers besides the rate value)
        if last_rate:
            last_cleaned = re.sub(r'(?:rs|re)[.,\s]*\d+', '', last_line, flags=re.I).strip()
            last_nums = re.findall(r'\b\d{2,5}\b', last_cleaned)
            if not last_nums:
                # Pre-set the rate in context
                ctx.rate = last_rate

    all_entries = []
    all_traces: List[TokenTrace] = []
    normalized_lines = []

    # Pre-scan for trailing "Each-N" line (standalone qty modifier for entire message)
    trailing_each_qty = None
    if non_empty_lines:
        last_l = non_empty_lines[-1].strip()
        each_m = re.match(r'^(?:each|ecsh|ech|eash|ea)\s*[-.]?\s*(\d+)\s*(?:set|sat|ser|seat|pcs|pes)?$', last_l, re.I)
        if each_m:
            trailing_each_qty = int(each_m.group(1))

    for line in lines:
        line = line.strip()
        if not line:
            continue
        line_traces: List[TokenTrace] = []
        entries = parse_line(line, ctx, message_id, meta, stats, line_traces)
        all_entries.extend(entries)
        all_traces.extend(line_traces)
        normalized_lines.append(normalize_separators(line))

    # Apply trailing "Each-N" retroactively to all entries with qty=1
    if trailing_each_qty and all_entries:
        for entry in all_entries:
            if entry.qty == 1:
                entry.qty = trailing_each_qty
                entry.amount = (entry.rate or 0) * trailing_each_qty

    # Timeslot fallback: if still None after parsing + inheritance, derive from timestamp
    if ctx.timeslot is None and ctx.lottery and timestamp:
        inferred_ts = resolve_timeslot_from_timestamp(timestamp, ctx.lottery)
        if inferred_ts:
            ctx.timeslot = inferred_ts
            # Backfill entries that were created with timeslot=None
            for entry in all_entries:
                if entry.timeslot is None:
                    entry.timeslot = inferred_ts

    # Contest date: only set when explicitly mentioned in message text
    if ctx.date_str:
        for entry in all_entries:
            entry.contest_date = ctx.date_str

    stats["messages_with_entries"] += (1 if all_entries else 0)
    stats["messages_without_entries"] += (1 if not all_entries else 0)

    # Build message trace for audit sheet
    unknown_count = sum(1 for t in all_traces if t.interpretation.startswith("UNKNOWN:"))
    total_tokens = len(all_traces) if all_traces else 1
    bet_count = sum(1 for t in all_traces if t.interpretation.startswith("BET_NUMBER:"))

    # Confidence scoring
    if not all_entries:
        # No entries extracted — low if message has numbers, otherwise it's just noise
        has_numbers = bool(re.search(r'\d{2,5}', text))
        confidence = "low" if has_numbers else "medium"
    elif total_tokens > 0:
        unknown_ratio = unknown_count / total_tokens
        if unknown_ratio == 0:
            confidence = "high"
        elif unknown_ratio < 0.15:
            confidence = "high"
        elif unknown_ratio < 0.35:
            confidence = "medium"
        else:
            confidence = "low"
    else:
        confidence = "high"

    msg_trace = MessageTrace(
        message_id=message_id,
        raw_text=text,
        normalized_text=" | ".join(normalized_lines),
        token_traces=all_traces,
        entries_count=len(all_entries),
        lottery=ctx.lottery,
        timeslot=ctx.timeslot,
        rate=ctx.rate,
        confidence=confidence,
        group_name=group_name,
        sender=message.get("sender", ""),
        push_name=message.get("push_name", ""),
        timestamp=timestamp,
        human_time=datetime.fromtimestamp(timestamp, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if timestamp else "",
    )

    # Collect traces in stats
    if "message_traces" not in stats:
        stats["message_traces"] = []
    stats["message_traces"].append(msg_trace)

    # Return the timeslot found in this message (or inherited) for next message
    return all_entries, ctx.timeslot

# ============================================================
# LOAD JSONL FILES
# ============================================================

def load_jsonl_files(input_dir: str) -> Dict[str, List[dict]]:
    """Load all JSONL files from directory, grouped by filename."""
    messages_by_file = {}
    files = sorted(glob.glob(os.path.join(input_dir, "*_messages.jsonl")))

    if not files:
        logger.warning(f"No JSONL files found in {input_dir}")
        return messages_by_file

    for filepath in files:
        filename = os.path.basename(filepath)
        messages = []
        with open(filepath, "r", encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    obj["_source_file"] = filename
                    obj["_source_line"] = line_num
                    messages.append(obj)
                except json.JSONDecodeError as e:
                    logger.warning(f"JSON error in {filename}:{line_num}: {e}")
        messages_by_file[filename] = messages
        logger.info(f"Loaded {len(messages)} messages from {filename}")

    return messages_by_file

# ============================================================
# EXCEL REPORT GENERATOR (requires pandas + openpyxl)
# ============================================================

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="2E75B6")
SUMMARY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
NUMBER_FORMAT_CURRENCY = '#,##0'


def style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:500]:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def create_dashboard(wb, all_entries_df, stats):
    """Create a summary dashboard as the first sheet."""
    ws = wb.active
    ws.title = "Dashboard"

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "LOTTERY TICKET REPORT — DASHBOARD"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Parser V{VERSION}"
    ws['A2'].font = Font(italic=True, color="666666")
    ws.merge_cells('A2:H2')

    row = 4

    # --- Overall Stats ---
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "PROCESSING SUMMARY"
    ws[f'A{row}'].font = SUBTITLE_FONT
    row += 1

    stat_items = [
        ("Total Messages Processed", stats.get("total_messages", 0)),
        ("Messages with Entries", stats.get("messages_with_entries", 0)),
        ("Messages Skipped (Noise)", stats.get("skipped_noise", 0)),
        ("Total Entries Extracted", stats.get("extracted", 0)),
        ("Unknown Tokens", len(stats.get("unknown_tokens", {}))),
    ]
    for label, value in stat_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = NUMBER_FORMAT_CURRENCY
        row += 1

    if len(all_entries_df) == 0:
        ws[f'A{row+1}'] = "No entries extracted."
        return

    row += 1

    # --- Tickets by Lottery Type ---
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "TICKETS BY LOTTERY TYPE"
    ws[f'A{row}'].font = SUBTITLE_FONT
    row += 1

    headers = ["Lottery", "Ticket Count", "Total Qty", "Total Amount (Rs)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    lottery_summary = all_entries_df.groupby("lottery").agg(
        ticket_count=("number", "count"),
        total_qty=("qty", "sum"),
        total_amount=("amount", "sum")
    ).reset_index()

    for _, r in lottery_summary.iterrows():
        ws.cell(row=row, column=1, value=r["lottery"] or "UNKNOWN")
        ws.cell(row=row, column=2, value=int(r["ticket_count"]))
        ws.cell(row=row, column=3, value=int(r["total_qty"]))
        ws.cell(row=row, column=4, value=int(r["total_amount"]))
        ws[f'D{row}'].number_format = NUMBER_FORMAT_CURRENCY
        row += 1

    # Grand total
    ws.cell(row=row, column=1, value="TOTAL")
    ws[f'A{row}'].font = Font(bold=True)
    ws.cell(row=row, column=2, value=int(lottery_summary["ticket_count"].sum()))
    ws.cell(row=row, column=3, value=int(lottery_summary["total_qty"].sum()))
    ws.cell(row=row, column=4, value=int(lottery_summary["total_amount"].sum()))
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).font = Font(bold=True)
    row += 2

    # --- Tickets by Category ---
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "TICKETS BY CATEGORY"
    ws[f'A{row}'].font = SUBTITLE_FONT
    row += 1

    headers = ["Category", "Ticket Count", "Total Qty", "Total Amount (Rs)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    cat_summary = all_entries_df.groupby("category").agg(
        ticket_count=("number", "count"),
        total_qty=("qty", "sum"),
        total_amount=("amount", "sum")
    ).reset_index().sort_values("total_amount", ascending=False)

    for _, r in cat_summary.iterrows():
        ws.cell(row=row, column=1, value=r["category"])
        ws.cell(row=row, column=2, value=int(r["ticket_count"]))
        ws.cell(row=row, column=3, value=int(r["total_qty"]))
        ws.cell(row=row, column=4, value=int(r["total_amount"]))
        ws[f'D{row}'].number_format = NUMBER_FORMAT_CURRENCY
        row += 1

    row += 1

    # --- Tickets by Lottery + Timeslot ---
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "TICKETS BY LOTTERY × TIMESLOT"
    ws[f'A{row}'].font = SUBTITLE_FONT
    row += 1

    headers = ["Lottery", "Timeslot", "Ticket Count", "Total Qty", "Total Amount (Rs)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    lt_summary = all_entries_df.groupby(["lottery", "timeslot"]).agg(
        ticket_count=("number", "count"),
        total_qty=("qty", "sum"),
        total_amount=("amount", "sum")
    ).reset_index()

    for _, r in lt_summary.iterrows():
        ws.cell(row=row, column=1, value=r["lottery"] or "UNKNOWN")
        ws.cell(row=row, column=2, value=r["timeslot"] or "UNKNOWN")
        ws.cell(row=row, column=3, value=int(r["ticket_count"]))
        ws.cell(row=row, column=4, value=int(r["total_qty"]))
        ws.cell(row=row, column=5, value=int(r["total_amount"]))
        ws[f'E{row}'].number_format = NUMBER_FORMAT_CURRENCY
        row += 1

    row += 1

    # --- Tickets by Bet Type ---
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "TICKETS BY BET TYPE"
    ws[f'A{row}'].font = SUBTITLE_FONT
    row += 1

    headers = ["Bet Type", "Ticket Count", "Total Qty", "Total Amount (Rs)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    bt_summary = all_entries_df.groupby("bet_type").agg(
        ticket_count=("number", "count"),
        total_qty=("qty", "sum"),
        total_amount=("amount", "sum")
    ).reset_index().sort_values("total_amount", ascending=False)

    for _, r in bt_summary.iterrows():
        ws.cell(row=row, column=1, value=r["bet_type"] or "DIRECT")
        ws.cell(row=row, column=2, value=int(r["ticket_count"]))
        ws.cell(row=row, column=3, value=int(r["total_qty"]))
        ws.cell(row=row, column=4, value=int(r["total_amount"]))
        ws[f'D{row}'].number_format = NUMBER_FORMAT_CURRENCY
        row += 1

    auto_width(ws, 45)


def create_entries_sheet(wb, entries_df):
    """Create the parsed entries data sheet."""
    ws = wb.create_sheet("Parsed_Entries")

    columns = [
        "message_id", "source_file", "human_time", "group_name", "push_name",
        "lottery", "timeslot", "bet_type", "number", "category",
        "qty", "rate", "amount", "is_box", "raw_line"
    ]

    # Headers
    for c, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=c, value=col_name)
    style_header_row(ws, 1, len(columns))

    # Data
    for idx, (_, row) in enumerate(entries_df.iterrows(), 2):
        for c, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            if isinstance(val, bool):
                val = "Yes" if val else ""
            ws.cell(row=idx, column=c, value=val)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws)


def create_raw_messages_sheet(wb, messages):
    """Create raw messages reference sheet."""
    ws = wb.create_sheet("Raw_Messages")

    columns = ["message_id", "timestamp", "human_time", "group_name", "push_name", "text", "source_file"]
    for c, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=c, value=col_name)
    style_header_row(ws, 1, len(columns))

    row_num = 2
    for msg in messages:
        text = msg.get("text", "")
        if not text:
            continue
        ts = int(msg.get("whatsapp_timestamp", 0))
        human_time = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S") if ts else ""
        ws.cell(row=row_num, column=1, value=msg.get("message_id", ""))
        ws.cell(row=row_num, column=2, value=ts)
        ws.cell(row=row_num, column=3, value=human_time)
        ws.cell(row=row_num, column=4, value=msg.get("group_name", ""))
        ws.cell(row=row_num, column=5, value=msg.get("push_name", ""))
        ws.cell(row=row_num, column=6, value=text)
        ws.cell(row=row_num, column=7, value=msg.get("_source_file", ""))
        ws.cell(row=row_num, column=6).alignment = Alignment(wrap_text=True)
        row_num += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws, 60)


def create_unknown_tokens_sheet(wb, unknown_tokens):
    """Sheet listing unknown tokens for review."""
    ws = wb.create_sheet("Unknown_Tokens")

    ws.cell(row=1, column=1, value="Token")
    ws.cell(row=1, column=2, value="Count")
    style_header_row(ws, 1, 2)

    sorted_tokens = sorted(unknown_tokens.items(), key=lambda x: x[1], reverse=True)
    for idx, (token, count) in enumerate(sorted_tokens, 2):
        ws.cell(row=idx, column=1, value=token)
        ws.cell(row=idx, column=2, value=count)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws)


CONF_FILL_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # Green
CONF_FILL_MEDIUM = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Orange/Yellow
CONF_FILL_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Red/Pink

CONF_FONT_HIGH = Font(color="006100")
CONF_FONT_MEDIUM = Font(color="9C6500")
CONF_FONT_LOW = Font(color="9C0006")


def create_parse_audit_sheet(wb, stats):
    """Create the Parse Audit sheet showing raw → interpretation for each message."""
    traces = stats.get("message_traces", [])
    if not traces:
        return

    ws = wb.create_sheet("Parse_Audit")

    # Headers
    headers = [
        "Timestamp", "Group", "Sender", "Confidence",
        "Raw Message", "Normalized", "Token Breakdown",
        "Entries", "Lottery", "Timeslot", "Rate"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    for mt in traces:
        # Build token breakdown string: "dr→LOTTERY:DEAR | 1pm→TIMESLOT:1PM | 360→BET_NUMBER:360 | *5→QTY:5"
        breakdown_parts = []
        for tt in mt.token_traces:
            breakdown_parts.append(f"{tt.original}→{tt.interpretation}")
        breakdown = " | ".join(breakdown_parts) if breakdown_parts else ""

        ws.cell(row=row, column=1, value=mt.human_time)
        ws.cell(row=row, column=2, value=mt.group_name)
        ws.cell(row=row, column=3, value=mt.push_name or mt.sender)
        ws.cell(row=row, column=4, value=mt.confidence.upper())
        ws.cell(row=row, column=5, value=mt.raw_text)
        ws.cell(row=row, column=6, value=mt.normalized_text)
        ws.cell(row=row, column=7, value=breakdown)
        ws.cell(row=row, column=8, value=mt.entries_count)
        ws.cell(row=row, column=9, value=mt.lottery or "")
        ws.cell(row=row, column=10, value=mt.timeslot or "")
        ws.cell(row=row, column=11, value=f"Rs{mt.rate}" if mt.rate else "")

        # Color the entire row by confidence
        if mt.confidence == "high":
            fill, font = CONF_FILL_HIGH, CONF_FONT_HIGH
        elif mt.confidence == "medium":
            fill, font = CONF_FILL_MEDIUM, CONF_FONT_MEDIUM
        else:
            fill, font = CONF_FILL_LOW, CONF_FONT_LOW

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = THIN_BORDER

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 20   # Timestamp
    ws.column_dimensions['B'].width = 30   # Group
    ws.column_dimensions['C'].width = 15   # Sender
    ws.column_dimensions['D'].width = 12   # Confidence
    ws.column_dimensions['E'].width = 50   # Raw Message
    ws.column_dimensions['F'].width = 50   # Normalized
    ws.column_dimensions['G'].width = 80   # Token Breakdown
    ws.column_dimensions['H'].width = 10   # Entries
    ws.column_dimensions['I'].width = 10   # Lottery
    ws.column_dimensions['J'].width = 10   # Timeslot
    ws.column_dimensions['K'].width = 10   # Rate


def generate_excel_report(
    output_path: str,
    all_entries: List[CanonicalEntry],
    all_messages: List[dict],
    stats: dict
):
    """Generate the complete Excel report."""
    wb = Workbook()

    entries_df = pd.DataFrame([asdict(e) for e in all_entries]) if all_entries else pd.DataFrame()

    # Fill NaN
    if len(entries_df) > 0:
        entries_df["lottery"] = entries_df["lottery"].fillna("UNKNOWN")
        entries_df["timeslot"] = entries_df["timeslot"].fillna("UNKNOWN")
        entries_df["bet_type"] = entries_df["bet_type"].fillna("DIRECT")

    create_dashboard(wb, entries_df, stats)
    create_entries_sheet(wb, entries_df)
    create_parse_audit_sheet(wb, stats)
    create_raw_messages_sheet(wb, all_messages)
    create_unknown_tokens_sheet(wb, stats.get("unknown_tokens", {}))

    wb.save(output_path)
    logger.info(f"Report saved: {output_path}")


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Lottery Message Parser V5")
    parser.add_argument("--input", default="data/raw", help="Input directory with JSONL files")
    parser.add_argument("--output", default="reports", help="Output directory for Excel reports")
    args = parser.parse_args()

    input_dir = args.input
    output_dir = args.output
    os.makedirs(output_dir, exist_ok=True)

    logger.info(f"=== Lottery Parser V{VERSION} Starting ===")
    logger.info(f"Input: {input_dir} | Output: {output_dir}")

    messages_by_file = load_jsonl_files(input_dir)

    if not messages_by_file:
        logger.error("No files to process.")
        return

    # --- Process all files together for combined report ---
    all_entries = []
    all_messages = []
    stats = {
        "total_messages": 0,
        "messages_with_entries": 0,
        "messages_without_entries": 0,
        "skipped_noise": 0,
        "extracted": 0,
        "unknown_tokens": {},
    }

    for filename, messages in messages_by_file.items():
        logger.info(f"Processing {filename}: {len(messages)} messages")
        stats["total_messages"] += len(messages)

        prev_timeslot = None
        prev_group = None
        for msg in messages:
            all_messages.append(msg)
            entries, prev_timeslot = parse_message(msg, stats, prev_timeslot, prev_group)
            prev_group = msg.get("group_name", "")
            all_entries.extend(entries)

    logger.info(f"Total entries extracted: {len(all_entries)}")
    logger.info(f"Messages with entries: {stats['messages_with_entries']}")
    logger.info(f"Noise skipped: {stats['skipped_noise']}")
    logger.info(f"Unknown token types: {len(stats['unknown_tokens'])}")

    # --- Generate combined report ---
    combined_output = os.path.join(output_dir, "lottery_report_combined.xlsx")
    generate_excel_report(combined_output, all_entries, all_messages, stats)

    # --- Also generate per-date reports ---
    for filename, messages in messages_by_file.items():
        date_prefix = filename.replace("_messages.jsonl", "")
        per_date_stats = {
            "total_messages": len(messages),
            "messages_with_entries": 0,
            "messages_without_entries": 0,
            "skipped_noise": 0,
            "extracted": 0,
            "unknown_tokens": {},
        }
        file_entries = []
        prev_timeslot_pd = None
        prev_group_pd = None
        for msg in messages:
            entries, prev_timeslot_pd = parse_message(msg, per_date_stats, prev_timeslot_pd, prev_group_pd)
            prev_group_pd = msg.get("group_name", "")
            file_entries.extend(entries)

        if file_entries:
            per_date_output = os.path.join(output_dir, f"{date_prefix}_lottery_report.xlsx")
            generate_excel_report(per_date_output, file_entries, messages, per_date_stats)

    # --- Print top unknown tokens ---
    top_unknown = sorted(stats["unknown_tokens"].items(), key=lambda x: x[1], reverse=True)[:20]
    if top_unknown:
        logger.info("Top 20 unknown tokens:")
        for token, count in top_unknown:
            logger.info(f"  {token}: {count}")

    logger.info(f"=== Parser V{VERSION} Complete ===")


if __name__ == "__main__":
    main()
